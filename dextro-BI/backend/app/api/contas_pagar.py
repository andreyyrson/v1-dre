import io
from datetime import date
from enum import Enum
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from app.dependencies import get_contas_service
from app.exceptions import PeriodoInvalidoError
from app.models.conta_pagar import ContaPagar, ContasPagarResponse, Paginacao
from app.services.contas_service import ContasService
from app.auth.deps import get_current_user
from app.models.user import User

router = APIRouter(tags=["contas-pagar"])


class SortBy(str, Enum):
    data_vencimento = "data_vencimento"
    valor = "valor"
    fornecedor = "fornecedor"


class Order(str, Enum):
    asc = "asc"
    desc = "desc"


async def _buscar_contas(
    service: ContasService,
    id_empresa: int,
    data_inicial: date,
    data_final: date,
) -> list[ContaPagar]:
    try:
        return await service.listar_contas(
            data_inicial=data_inicial,
            data_final=data_final,
            id_empresa=id_empresa,
        )
    except PeriodoInvalidoError as exc:
        raise HTTPException(status_code=422, detail=str(exc))


@router.get("/contas-pagar", response_model=ContasPagarResponse, response_model_by_alias=False)
async def listar_contas_pagar(
    id_empresa: int = Query(..., description="ID da empresa (entity)"),
    data_inicial: date = Query(..., description="Data inicial de vencimento (aaaa-mm-dd)"),
    data_final: date = Query(..., description="Data final de vencimento (aaaa-mm-dd)"),
    apenas_abertas: bool = Query(True, description="Retornar somente contas em aberto"),
    page: int = Query(1, ge=1, description="Página atual (>= 1)"),
    page_size: int = Query(50, ge=1, le=200, description="Itens por página (1-200)"),
    sort_by: SortBy = Query(SortBy.data_vencimento, description="Campo de ordenação"),
    order: Order = Query(Order.asc, description="Direção da ordenação"),
    service: ContasService = Depends(get_contas_service),
    _: User = Depends(get_current_user),
) -> ContasPagarResponse:
    todas = await _buscar_contas(service, id_empresa, data_inicial, data_final)

    # Resumo (KPIs) calculado sobre TODAS as contas (pagas + em aberto).
    resumo = service.calcular_resumo(todas)

    # Filtra a lista se solicitado apenas abertas.
    contas = [c for c in todas if not c.pago] if apenas_abertas else todas

    page_items, total, total_pages = service.ordenar_e_paginar(
        contas, sort_by=sort_by.value, order=order.value, page=page, page_size=page_size
    )
    return ContasPagarResponse(
        contas=page_items,
        resumo=resumo,
        paginacao=Paginacao(
            page=page, page_size=page_size, total=total, total_pages=total_pages
        ),
    )


@router.post("/contas-pagar/refresh")
async def refresh_contas_pagar(
    id_empresa: int = Query(..., description="ID da empresa (entity)"),
    data_inicial: date = Query(..., description="Data inicial de vencimento (aaaa-mm-dd)"),
    data_final: date = Query(..., description="Data final de vencimento (aaaa-mm-dd)"),
    service: ContasService = Depends(get_contas_service),
    _: User = Depends(get_current_user),
) -> dict:
    """Invalida o cache e força uma nova busca na API BomControle na próxima requisição."""
    service.invalidar_cache(id_empresa, data_inicial, data_final)
    return {"message": "Cache invalidado"}


@router.get("/contas-pagar/export")
async def exportar_contas_pagar(
    id_empresa: int = Query(..., description="ID da empresa (entity)"),
    data_inicial: date = Query(..., description="Data inicial de vencimento (aaaa-mm-dd)"),
    data_final: date = Query(..., description="Data final de vencimento (aaaa-mm-dd)"),
    apenas_abertas: bool = Query(True, description="Considerar somente contas em aberto"),
    sort_by: SortBy = Query(SortBy.data_vencimento, description="Campo de ordenação"),
    order: Order = Query(Order.asc, description="Direção da ordenação"),
    ids: Optional[str] = Query(
        None,
        description="IDs (GUID) separados por vírgula. Se omitido, exporta toda a busca atual.",
    ),
    service: ContasService = Depends(get_contas_service),
    _: User = Depends(get_current_user),
) -> StreamingResponse:
    """Exporta as contas em Excel (XLSX). Sem `ids`: toda a busca; com `ids`: só as selecionadas."""
    todas = await _buscar_contas(service, id_empresa, data_inicial, data_final)
    contas = [c for c in todas if not c.pago] if apenas_abertas else todas

    if ids:
        selecionados = [i.strip() for i in ids.split(",") if i.strip()]
        contas = service.filtrar_selecionadas(contas, selecionados)

    contas = service.ordenar(contas, sort_by=sort_by.value, order=order.value)

    wb = Workbook()
    ws = wb.active
    ws.title = "Contas a Pagar"

    # Headers
    headers = ["ID", "Descrição", "Fornecedor", "Categoria", "Valor", "Vencimento", "Quitação", "Status", "Empresa"]
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Data
    for row_num, c in enumerate(contas, 2):
        ws.cell(row=row_num, column=1, value=c.id)
        ws.cell(row=row_num, column=2, value=c.descricao or "")
        ws.cell(row=row_num, column=3, value=c.fornecedor or "")
        ws.cell(row=row_num, column=4, value=c.categoria or "")
        ws.cell(row=row_num, column=5, value=c.valor)
        ws.cell(row=row_num, column=6, value=c.data_vencimento.isoformat() if c.data_vencimento else "")
        ws.cell(row=row_num, column=7, value=c.data_quitacao.isoformat() if c.data_quitacao else "")
        ws.cell(row=row_num, column=8, value="Pago" if c.pago else "Em aberto")
        ws.cell(row=row_num, column=9, value=c.nome_empresa or "")

    # Auto-width columns
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"contas_pagar_{id_empresa}_{data_inicial}_{data_final}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
